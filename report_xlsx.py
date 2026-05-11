import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment # Importar Alignment

def build_evidence_xlsx(inp_dict: dict, results: dict) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Evidencia"

    ws["A1"] = "EVIDENCIA – Cálculo Burden / Selección TC"
    ws["A1"].font = Font(bold=True, size=14)

    # Inputs
    ws["A3"] = "Entradas"
    ws["A3"].font = Font(bold=True, color="0000FF")
    row = 4
    for k, v in inp_dict.items():
        ws[f"A{row}"] = k
        ws[f"B{row}"] = str(v)
        row += 1

    # Results Summary
    row += 1
    ws[f"A{row}"] = "Resultados"
    ws[f"A{row}"].font = Font(bold=True)
    row += 1

    ws[f"A{row}"] = "kVA Autorizados (Calculado)"
    ws[f"B{row}"] = results.get("kva_autorizado_calc")
    row += 1

    ws[f"A{row}"] = "kVA Restantes"
    ws[f"B{row}"] = results.get("kva_restantes")
    row += 1

    ws[f"A{row}"] = "RESULTADO FINAL"
    ws[f"B{row}"] = results["resultado"]
    ws[f"B{row}"].font = Font(bold=True, color="008000" if results["resultado"] == "cumple" else "FF0000")
    row += 2

    # Detalle por fase en tabla
    ws[f"A{row}"] = "Detalle por Fase"
    ws[f"A{row}"].font = Font(bold=True)
    row += 1
    
    headers = ["Fase", "Relación", "Serie", "Marca", "VA TC", "Burden Total", "Utilización", "Cumple"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    
    row += 1
    for fase in results["detalle_fases"]:
        for col, key in enumerate(["fase", "relacion", "serie", "marca", "va_tc", "burden_total", "utilizacion", "cumple"], 1):
            ws.cell(row=row, column=col, value=fase[key])
        row += 1

    # Simple style
    for c in ["A", "B"]:
        ws.column_dimensions[c].width = 45

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
