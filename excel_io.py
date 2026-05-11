import io
from openpyxl import load_workbook

def read_parametros(xlsx_bytes: bytes) -> dict:
    wb = load_workbook(filename=bytes_to_filelike(xlsx_bytes), data_only=True)
    
    # Búsqueda robusta de la hoja "Parametros"
    sheet_name = None
    for name in wb.sheetnames:
        if name.strip().lower() == "parametros":
            sheet_name = name
            break
    
    # Si no se encuentra por nombre, usamos la hoja activa como respaldo
    ws = wb[sheet_name] if sheet_name else wb.active

    # Mapeo etiqueta (col A) -> valor (col B)
    data = {}
    for r in range(1, ws.max_row + 1):
        key = ws.cell(r, 1).value
        val = ws.cell(r, 2).value
        if isinstance(key, str) and key.strip():
            data[key.strip()] = val
    return data

def bytes_to_filelike(b: bytes):
    return io.BytesIO(b)
